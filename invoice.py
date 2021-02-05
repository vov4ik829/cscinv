from shutil import copyfile
from os import path
from openpyxl import load_workbook
from workbook import load_workbook as load_ro_workbook
from textlogger import textlogger


from report_readers import customs_report

class Invoice():
    #Fulls 40/20 price is 50
    #Empties 40 price is 40
    #Empties 20 price is 20, but empty 20s are shipped in couples, odd remainder should be
    #considered as two for price calculation purposes
    def __init__(self, record, inv_date:str):
        self.car_id = record[3]
        self.total = 0
        self.total_empty_twenties = 0
        self.records = []
        self.odd_empty_twenties=False
        self.date = inv_date
        self.add_record(record)

    def add_record(self, record):
        # Record -  list [container_id, type, full/empty, car_id]
        assert record[1][0:2] in ['20','40', '45']
        assert record[2] in ['full', 'empty']
        assert record[3] == self.car_id
        if record[2] == 'full':
            self.add_to_total(50)
        else:
            if record[1][0:2]=='40':
                self.add_to_total(40)

            if record[1][0:2]=='45':
                self.add_to_total(40)

            if record[1][0:2]=='20':
                self.add_to_total_empty_twenties()

        self.records.append(record)
        textlogger.log('Added '+' '.join(record))

    def add_to_total(self, amm):
        assert amm in [40,50]
        self.total+=amm

    def add_to_total_empty_twenties(self):
        self.odd_empty_twenties = not self.odd_empty_twenties
        if self.odd_empty_twenties:
            self.total_empty_twenties+=40

    def get_total(self):
        return self.total+self.total_empty_twenties
        
    def get_records(self):

        return self.records

class InvoiceWriter():
    def __init__(self, driver_list:str , template:str):
        driver_col = 'B'
        id_col = 'C'
        account_col = 'D'
        car_id_col = 'F'
        bank_col = 'E'
        textlogger.log("Started reading drivers list ")
        wb = load_ro_workbook(driver_list)
        # TODO duplicate car_id checking
        ws = wb.worksheets[0]
        cars ={}
        row = 2
        while ws['{}{}'.format(driver_col, row)].value is not None:
            driver_name = ws['{}{}'.format(driver_col, row)].value
            driver_id = ws['{}{}'.format(id_col, row)].value
            driver_account = ws['{}{}'.format(account_col, row)].value
            car_id = ws['{}{}'.format(car_id_col, row)].value
            # Duplicate car check code should go here
            bank = ws['{}{}'.format(bank_col, row)].value
            cars.update({car_id:{'name':driver_name, 
                                'id':driver_id, 
                                'account': driver_account,
                                'bank': bank}})
            row+=1
        wb.close()
        textlogger.log('Finished reading drivers list')
        self.cars = cars
        self.template = template
        
    def write_invoice(self, invoice:Invoice, dst_folder):
        try:
            car = self.cars[invoice.car_id] # Handle Key Error
        except KeyError:
            textlogger.log('Car {} not found in drivers list. Skipping'.format(invoice.car_id))
            return
        car_id = invoice.car_id
        driver_name = car['name']
        out_file_name = '_'.join([invoice.date, driver_name, car_id])+'.xlsx'
        out_file_name = path.join(dst_folder, out_file_name)
        copyfile(self.template, out_file_name)
        wb = load_workbook(out_file_name)
        ws = wb.worksheets[0]
        ws['A1'].value = ws['A1'].value.replace('saxeli_gvari', driver_name)
        ws['A2'].value = ws['A2'].value.replace('id', str(car['id']))
        ws['A7'].value = invoice.date
        # Record -  list [[container_id, type, full/empty, car_id],]
        cont_list = ', '.join([record[0] for record in invoice.get_records()])
        trans = {'full':'სავსე', 'empty':'ცარიელი'}
        sum_up = []
        for record in invoice.records:
            sum_up.append('{} {}'.format(record[1], trans[record[2]]))
        sum_up = ["{}X{}".format(sum_up.count(record), record)for record in sum_up]
        sum_up = list(set(sum_up))
        sum_up.sort(key=lambda x: x[7]+x[2])
        print(sum_up)
        ws['A13'].value = '{};\n{}'.format(cont_list, ', '.join(sum_up))
        ws['C26'].value = invoice.get_total()
        try:
            ws['C30'].value = car['bank']
        except KeyError:
            ws['C30'].value = car['account'][5:7]

        ws['C32'].value = car['account']
        ws['E46'].value = '{}. {}'.format(car['name'].split(' ')[0][0], car['name'].split(' ')[1])
        wb.save(out_file_name)
        textlogger.log('Processed invoice {}'.format(out_file_name))
        wb.close()
        return out_file_name

def process(source_file:str, dst_folder:str, invoice_date:str):
    # TODO Add exception logging
    try:
        records = customs_report(source_file)
    except:
        textlogger.log('Problem with report. Please check {}'.format(source_file))
        return
    invoices = {}
    try:
        for record in records:
            try:
                invoices[record[3]].add_record(record)
            except KeyError:
                invoices.update({record[3]:Invoice(record, invoice_date)})
    except:
        textlogger.log('Problem with report. Please check {}'.format(source_file))
        return
    try:
        writer = InvoiceWriter(path.join('input_files','BAZA.xlsx'), 
            path.join('input_files','invoice_template.xlsx'))
    except:
        textlogger.log('Problem with drivers list. Please check BAZA.xlsx')
        return
    count=0
    fault_invoices=[]
    for invoice in invoices.values():
        if writer.write_invoice(invoice, dst_folder):
           count+=1
        else:
            fault_invoices.append(invoice.car_id)
    textlogger.log('Completed. {} invoices written'.format(count))
    if fault_invoices != []:
        textlogger.log('Could not write invoices for cars: \n{}'.format(' '.join(fault_invoices)))