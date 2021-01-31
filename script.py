from shutil import copyfile
from os import path

from openpyxl.xml.constants import PACKAGE_CHARTS
from workbook import load_workbook as load_ro_workbook
from openpyxl import load_workbook

class Invoice():
    #Fulls 40/20 price is 50
    #Empties 40 price is 40
    #Empties 20 price is 20, but emptie 20s are shipped in couples, odd remainder should be
    #considered as two for price calculation means
    def __init__(self, record, inv_date:str):
        self.car_id = record[3]
        self.total = 0
        self.total_twenties = 0
        self.records = []
        self.empty_twenties=0
        self.odd_twenties=False
        self.date = inv_date
        self.add_record(record)

    def add_record(self, record):
        # Record -  list [container_id, type, full/empty, car_id]
        assert record[1][0:2] in [20,40]
        assert record[2] in ['full', 'empty']
        assert record[3] == self.car_id
        if record[2] == 'full':
            self.add_to_total(50)
        else:
            if record[1][0:2]==40:
                self.add_to_total(40)
            if record[1][0:2]==20:
                self.add_to_total_twenties

        self.records.append(record)

    def add_to_total(self, amm):
        assert amm in [40,50]
        self.total+=amm

    def add_to_total_twenties(self):
        self.odd_twenties = not self.odd_twenties
        if self.odd_twenties:
            self.total_twenties+=40

    def get_total(self):
        return self.total+self.total_twenties
        
    def get_records(self):

        return self.records

class InvoiceWriter():
    def __init__(self, driver_list:str , template:str):
        driver_col = 'B'
        id_col = 'C'
        account_col = 'D'
        car_id_col = 'E'
        wb = load_ro_workbook(driver_list)
        ws = wb.worksheets[0]
        cars ={}
        row = 2
        while ws['{}{}'.format(driver_col, row)].value is not None:
            driver_name = ws['{}{}'.format(driver_col, row).value]
            driver_id = ws['{}{}'.format(id_col, row).value]
            driver_account = ws['{}{}'.format(account_col, row).value]
            car_id = ws['{}{}'.format(car_id_col, row).value]
            cars.update({car_id:{'name':driver_name, 
                                'id':driver_id, 
                                'account': driver_account}})
            row+=1
        wb.close()
        self.cars = cars
        self.template = template
        
    def write_invoice(self, invoice:Invoice, dst_folder):
        car = self.cars[invoice.car_id]
        car_id = invoice.car_id
        driver_name = car['name']
        out_file_name = '_'.join(invoice.date, driver_name, car_id)
        out_file_name = path.join(dst_folder, out_file_name)
        copyfile(self.template, out_file_name)
        wb = load_workbook(out_file_name)
        ws = wb.worksheets[1]
        ws['A1'].value = ws['A1'].value.replace('saxeli_gvari', driver_name)
        ws['A2'].value = ws['A2'].value.replace('id', car['driver_id'])
        ws['A7'].value = invoice.date
        # Record -  list [container_id, type, full/empty, car_id]
        cont_list = ', '.join([record[0] for record in invoice.get_records()])
        trans = {'full':'სავსე', 'empty':'ცარიელი'}
        sum_up = []
        for record in invoice.records:
            sum_up+='{} {}'.format(record[1], trans[record[2]])
        sum_up = ["{}X{}".format(sum_up.count(record), record)for record in sum_up]
        sum_up = set(sorted(sum_up))
        ws['A13'].value = '{}; {}'.format(cont_list, sum_up)
        ws['C26'].value = invoice.get_total()
        banks = {
            'TB':"თი ბი სი ბანკი", "GB":'საქართველოს ბანკი'
            }#TODO add more banks
        try:
            ws['C30'].value = banks[car['account'][5:7]]
        except KeyError:
            ws['C30'].value = car['account'][5:7]

        ws['C32'].value = car['account']
        ws['E46'].value = '{}. {}'.format(car['name'].split(' ')[0][0], car['name'].split(' ')[1])
        wb.save()
        wb.close()