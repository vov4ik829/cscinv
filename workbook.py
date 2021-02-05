import xlrd
import openpyxl

# TODO: raise LastColumn/LastRow reached exception.


class CellOutOfRange(Exception):
    """Dummy Exception"""
    pass


def abc2index(abc_col: str):
    """Converts letter column notation to number"""
    base = 26
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    result = 0
    pos = 0
    for i in abc_col[::-1]:
        num = (alphabet.index(i)+1)*(base**pos)
        result += num
        pos += 1
    return result-1


def abc_colrow2rowcol(rowcol_abc: str):
    """Converts for example A1 to (1,1) """
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    numbers = '0123456789'
    if rowcol_abc[0] not in alphabet:
        raise TypeError('First symbol must be one of: {}'.format(alphabet))
    if rowcol_abc[-1] in alphabet:
        raise TypeError('The last symbol can not be one of: {}'.format(alphabet))
    col = ''
    row = ''
    for i in rowcol_abc:
        if i in alphabet:
            if len(row)>0:
                raise ValueError('Please check the value: {}'.format(rowcol_abc))
            col+=i
        if i in numbers:
            row+=i
    return((int(row)-1, col))


class Workbook():
    """Workbook with worksheets. Wrap around xlrd workbook"""

    def __init__(self, filename: str):

        self.__wb__ = xlrd.open_workbook(filename)
        self.worksheets = [Worksheet(ws) for ws in self.__wb__.sheets()]  # list of worksheets accessed by index
        self.sheetnames = self.__wb__.sheet_names()

    def __getitem__(self, item):
        # Returns worksheet by it's name
        ws = Worksheet(self.__wb__.sheet_by_name(item))
        return ws
    def close(self):
        pass


class Worksheet():
    """Excel worksheet supporting access by single cell as A1 and by range as A1:B2"""

    class Cell():
        """Excel cell"""
        def __init__(self, cell):
            self.__cell__ = cell
            if (self.__cell__.ctype == 0) or (self.__cell__.ctype == 6):
                self.value = None
            else:
                self.value = self.__cell__.value

    def __init__(self, ws: xlrd.sheet):
        self.__ws__ = ws

    def __getitem__(self, abc_colrow: str):
        """Get single cell value, or range of cells like A2:B2"""
        if ':' in abc_colrow:
            rows = []
            start, end = abc_colrow.split(':')
            start_rowcol = abc_colrow2rowcol(start)
            end_rowcol = abc_colrow2rowcol(end)
            start_row = start_rowcol[0]
            start_col = abc2index(start_rowcol[1])
            end_row = end_rowcol[0]
            end_col = abc2index(end_rowcol[1])
            row_c = start_row
            while row_c <= end_row:
                row = []
                col_c = start_col
                while col_c <= end_col:
                    try:
                        cell = self.Cell(self.__ws__.cell(row_c, col_c))
                        row.append(cell)
                        col_c+=1
                    except IndexError:
                        raise CellOutOfRange
                rows.append(row)
                row_c += 1
            return rows
        rowcol = abc_colrow2rowcol(abc_colrow)
        row = rowcol[0]
        col = abc2index(rowcol[1])
        try:
            cell = self.Cell(self.__ws__.cell(row, col))
        except IndexError:
            raise CellOutOfRange
        return cell


def load_workbook(filename: str):
    if filename[-3:].lower() == 'xls':
        wb = Workbook(filename)
        return wb
    elif filename[-4:].lower()=='xlsx':
        wb = openpyxl.load_workbook(filename) # readonly = True; Explain to Vakho
        def save(filename):
            # Explain to Vaxo
            pass
        wb.save = save

        return wb
    
    else:
        raise ValueError('Unknown file type {}'.format(filename))
